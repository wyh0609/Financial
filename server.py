# -*- coding: utf-8 -*-
import os
import traceback
import sys
import re
import socket

# 设置 DeepSeek API 配置
os.environ["DEEPSEEK_API_KEY"] = os.environ.get("DEEPSEEK_API_KEY", "")  # 请设置你的 API Key
os.environ["DEEPSEEK_BASE_URL"] = "https://api.deepseek.com/v1/chat/completions"
os.environ["DEEPSEEK_MODEL"] = "deepseek-chat"

from flask import Flask, request, jsonify, send_from_directory, send_file, Response, stream_with_context
from flask_cors import CORS
from werkzeug.serving import WSGIRequestHandler
from main import ChatFinance
from pdf_processor import process_uploaded_pdf
from pdf_generator import generate_analysis_pdf as _generate_analysis_pdf
from conversation_manager import conversation_manager

# 目录配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ALLPDF_DIR = os.path.join(BASE_DIR, "models", "allpdf")
ALLTXT_DIR = os.path.join(BASE_DIR, "models", "alltxt")

# 优先使用 Vue 构建产物，回退到原始 frontend 目录
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend-dist")
if not os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
    FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")

# 确保所有 JSON 响应使用 UTF-8 编码
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_MIMETYPE'] = 'application/json; charset=utf-8'
CORS(app)

bot = ChatFinance()
print(f"[DEBUG] LLM type: {type(bot.llm).__name__}")
print(f"[DEBUG] LLM model: {bot.llm.model}")
print(f"[DEBUG] LLM base_url: {bot.llm.base_url}")


@app.route("/")
def index():
    resp = send_from_directory(FRONTEND_DIR, "index.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.after_request
def add_no_cache_headers(response):
    if "static" in response.headers.get("Content-Type", "") or response.headers.get("Content-Type", "").startswith("text/"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.route("/api/chat", methods=["POST"])
def chat():
    try:
        data = request.get_json()
    except Exception as e:
        return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    if not data:
        return jsonify({"error": "请求数据为空"}), 400

    question = data.get("question", "").strip()
    conversation_id = data.get("conversation_id", "").strip()
    web_search = data.get("web_search", False)
    
    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    try:
        print(f"[API] 收到问题: {question} (对话ID: {conversation_id}, 联网搜索: {web_search})", flush=True)
        
        context = ""
        if conversation_id:
            context = conversation_manager.get_context_for_llm(conversation_id, max_messages=3)
            if context:
                question = f"以下是对话历史：\n{context}\n\n当前问题：{question}"
        
        # 使用 ask_stream 收集结果 + thinking_steps（供非H5端使用）
        result = {
            "answer": "",
            "intent": "",
            "entities": [],
            "file": "",
            "context": [],
            "is_open": False,
            "sources": [],
        }
        thinking_steps = []

        for event_type, event_data in bot.ask_stream(question, verbose=False, web_search=web_search):
            if event_type == "thinking_step":
                thinking_steps.append({
                    "step": event_data.get("step", 0),
                    "label": event_data.get("label", ""),
                    "status": event_data.get("status", ""),
                    "content": event_data.get("content", "")
                })
            elif event_type == "meta":
                result.update(event_data)
            elif event_type == "chunk":
                result["answer"] += event_data
            elif event_type == "done":
                result["answer"] = event_data.get("answer", result["answer"])
                for k in ("intent", "entities", "file", "context", "is_open", "sources"):
                    if event_data.get(k):
                        result[k] = event_data[k]

        result["thinking_steps"] = thinking_steps
        print(f"[API] 回答成功, thinking_steps={len(thinking_steps)}", flush=True)

        if conversation_id:
            conversation_manager.add_message(conversation_id, "user", data.get("question", ""))
            conversation_manager.add_message(
                conversation_id,
                "bot",
                result.get("answer", ""),
                extra_data={
                    "thinking": {
                        "intent": result.get("intent", ""),
                        "entities": result.get("entities", []),
                        "file": result.get("file", ""),
                        "context": result.get("context", []),
                        "is_open": result.get("is_open", False),
                    }
                }
            )

        return jsonify(result)
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        with open("error_log.txt", "a", encoding="utf-8") as f:
            f.write(f"\n=== ERROR at {__import__('datetime').datetime.now()} ===\n")
            f.write(f"Question: {question}\n")
            f.write(err_msg)
            f.write("\n")
        return jsonify({"error": str(e)}), 500


@app.route("/api/chat_stream", methods=["POST"])
def chat_stream():
    try:
        data = request.get_json()
    except Exception as e:
        return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    if not data:
        return jsonify({"error": "请求数据为空"}), 400

    question = data.get("question", "").strip()
    conversation_id = data.get("conversation_id", "").strip()

    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    def generate():
        try:
            print(f"[API-Stream] 收到问题: {question}", flush=True)

            # 关键：发送2KB SSE注释填充，强制Chrome立即开始处理流
            # Chrome会缓冲前1024字节，填充后立即开始推送
            yield ": " + " " * 2048 + "\n\n"

            context = ""
            if conversation_id:
                context = conversation_manager.get_context_for_llm(conversation_id, max_messages=3)
                if context:
                    question_with_ctx = f"以下是对话历史：\n{context}\n\n当前问题：{question}"
                else:
                    question_with_ctx = question
            else:
                question_with_ctx = question

            import json as _json
            import time

            for event_type, event_data in bot.ask_stream(question_with_ctx, verbose=False):
                if event_type == "thinking_step":
                    step_info = {
                        "type": "thinking_step",
                        "step": event_data.get("step", 0),
                        "label": event_data.get("label", ""),
                        "status": event_data.get("status", ""),
                        "content": event_data.get("content", "")
                    }
                    print(f"[SSE] thinking_step {step_info['step']} {step_info['status']}", flush=True)
                    yield f"data: {_json.dumps(step_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)  # 强制让出线程，让 Werkzeug flush
                elif event_type == "meta":
                    meta_info = {
                        "type": "meta",
                        "intent": event_data.get("intent", ""),
                        "entities": event_data.get("entities", []),
                        "file": event_data.get("file", ""),
                        "context": event_data.get("context", []),
                        "is_open": event_data.get("is_open", False),
                        "sources": event_data.get("sources", [])
                    }
                    print(f"[SSE] meta", flush=True)
                    yield f"data: {_json.dumps(meta_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "chunk":
                    chunk_info = {"type": "chunk", "content": event_data}
                    yield f"data: {_json.dumps(chunk_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "done":
                    done_info = {
                        "type": "done",
                        "answer": event_data.get("answer", ""),
                        "intent": event_data.get("intent", ""),
                        "entities": event_data.get("entities", []),
                        "file": event_data.get("file", ""),
                        "context": event_data.get("context", []),
                        "is_open": event_data.get("is_open", False),
                        "sources": event_data.get("sources", [])
                    }
                    print(f"[SSE] done", flush=True)
                    yield f"data: {_json.dumps(done_info, ensure_ascii=False)}\n\n"

                    if conversation_id:
                        conversation_manager.add_message(conversation_id, "user", question)
                        conversation_manager.add_message(
                            conversation_id,
                            "bot",
                            event_data.get("answer", ""),
                            extra_data={
                                "thinking": {
                                    "intent": event_data.get("intent", ""),
                                    "entities": event_data.get("entities", []),
                                    "file": event_data.get("file", ""),
                                    "context": event_data.get("context", []),
                                    "is_open": event_data.get("is_open", False),
                                }
                            }
                        )

        except Exception as e:
            import traceback as tb
            err_msg = tb.format_exc()
            print(err_msg, flush=True)
            import json as _json
            error_info = {"type": "error", "content": str(e)}
            yield f"data: {_json.dumps(error_info, ensure_ascii=False)}\n\n"

    resp = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )
    return resp


@app.route("/api/chat_stream_get", methods=["GET"])
def chat_stream_get():
    """GET 版本的流式聊天接口，供前端 EventSource 使用"""
    question = request.args.get("question", "").strip()
    conversation_id = request.args.get("conversation_id", "").strip()

    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    def generate():
        try:
            print(f"[API-Stream-GET] 收到问题: {question}", flush=True)

            yield ": " + " " * 2048 + "\n\n"

            context = ""
            if conversation_id:
                context = conversation_manager.get_context_for_llm(conversation_id, max_messages=3)
                if context:
                    question_with_ctx = f"以下是对话历史：\n{context}\n\n当前问题：{question}"
                else:
                    question_with_ctx = question
            else:
                question_with_ctx = question

            import json as _json
            import time

            for event_type, event_data in bot.ask_stream(question_with_ctx, verbose=False):
                if event_type == "thinking_step":
                    step_info = {
                        "type": "thinking_step",
                        "step": event_data.get("step", 0),
                        "label": event_data.get("label", ""),
                        "status": event_data.get("status", ""),
                        "content": event_data.get("content", "")
                    }
                    yield f"data: {_json.dumps(step_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "meta":
                    meta_info = {
                        "type": "meta",
                        "intent": event_data.get("intent", ""),
                        "entities": event_data.get("entities", []),
                        "file": event_data.get("file", ""),
                        "context": event_data.get("context", []),
                        "is_open": event_data.get("is_open", False),
                        "sources": event_data.get("sources", [])
                    }
                    yield f"data: {_json.dumps(meta_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "chunk":
                    chunk_info = {"type": "chunk", "content": event_data}
                    yield f"data: {_json.dumps(chunk_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "done":
                    done_info = {
                        "type": "done",
                        "answer": event_data.get("answer", ""),
                        "intent": event_data.get("intent", ""),
                        "entities": event_data.get("entities", []),
                        "file": event_data.get("file", ""),
                        "context": event_data.get("context", []),
                        "is_open": event_data.get("is_open", False),
                        "sources": event_data.get("sources", [])
                    }
                    yield f"data: {_json.dumps(done_info, ensure_ascii=False)}\n\n"

                    if conversation_id:
                        conversation_manager.add_message(conversation_id, "user", question)
                        conversation_manager.add_message(
                            conversation_id,
                            "bot",
                            event_data.get("answer", ""),
                            extra_data={
                                "thinking": {
                                    "intent": event_data.get("intent", ""),
                                    "entities": event_data.get("entities", []),
                                    "file": event_data.get("file", ""),
                                    "context": event_data.get("context", []),
                                    "is_open": event_data.get("is_open", False),
                                }
                            }
                        )

        except Exception as e:
            import traceback as tb
            err_msg = tb.format_exc()
            print(err_msg, flush=True)
            import json as _json
            error_info = {"type": "error", "content": str(e)}
            yield f"data: {_json.dumps(error_info, ensure_ascii=False)}\n\n"

    resp = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )
    return resp


@app.route("/api/upload_pdf", methods=["POST"])
def upload_pdf():
    print(f"[API] ===== upload_pdf 收到请求 =====", flush=True)
    print(f"[API] Content-Type: {request.content_type}", flush=True)
    print(f"[API] Content-Length: {request.content_length}", flush=True)
    print(f"[API] files keys: {list(request.files.keys())}", flush=True)
    print(f"[API] form keys: {list(request.form.keys())}", flush=True)
    try:
        if 'file' not in request.files:
            return jsonify({"error": "未找到文件"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "文件名为空"}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "只支持PDF文件"}), 400
        
        # 处理文件名编码（确保 UTF-8）
        filename = file.filename
        if isinstance(filename, bytes):
            filename = filename.decode('utf-8', errors='ignore')
        
        pdf_content = file.read()
        if len(pdf_content) == 0:
            return jsonify({"error": "文件内容为空"}), 400
        
        print(f"[API] 收到PDF上传: {filename} ({len(pdf_content)} bytes)", flush=True)
        
        result = process_uploaded_pdf(
            pdf_filename=filename,
            pdf_content=pdf_content,
            allpdf_dir=ALLPDF_DIR,
            alltxt_dir=ALLTXT_DIR
        )
        
        if result['success']:
            print(f"[API] PDF处理成功: {result['txt_filename']} ({result['row_count']} 行)", flush=True)
            try:
                bot.reload_reports()
                print(f"[API] 财报数据已重新索引", flush=True)
            except Exception as e:
                print(f"[API] 重新索引失败: {e}", flush=True)
            
            return jsonify({
                "success": True,
                "pdf_filename": result['pdf_filename'],
                "txt_filename": result['txt_filename'],
                "row_count": result['row_count'],
                "company": result['company'],
                "year": result['year']
            })
        else:
            print(f"[API] PDF处理失败: {result['error']}", flush=True)
            return jsonify({"error": result['error']}), 400
            
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload_pdf_base64", methods=["POST"])
def upload_pdf_base64():
    """App 端绕过 scoped storage：通过 FileReader 读文件内容，Base64 上传"""
    print(f"[API] ===== upload_pdf_base64 收到请求 =====", flush=True)
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "请求数据为空"}), 400
        
        filename = data.get("filename", "uploaded.pdf")
        content = data.get("content", "")
        if not content:
            return jsonify({"error": "文件内容为空"}), 400

        print(f"[API] Base64 文件名: {filename}, base64 长度: {len(content)}", flush=True)
        
        # 解码 base64
        import base64
        pdf_content = base64.b64decode(content)
        print(f"[API] Base64 解码成功: {len(pdf_content)} bytes", flush=True)
        
        if len(pdf_content) == 0:
            return jsonify({"error": "文件内容为空"}), 400

        # 调用 PDF 处理
        result = process_uploaded_pdf(
            pdf_filename=filename,
            pdf_content=pdf_content,
            allpdf_dir=ALLPDF_DIR,
            alltxt_dir=ALLTXT_DIR
        )
        
        if result['success']:
            print(f"[API] PDF处理成功: {result['txt_filename']} ({result['row_count']} 行)", flush=True)
            try:
                bot.reload_reports()
                print(f"[API] 财报数据已重新索引", flush=True)
            except Exception as e:
                print(f"[API] 重新索引失败: {e}", flush=True)
            
            return jsonify({
                "success": True,
                "pdf_filename": result['pdf_filename'],
                "txt_filename": result['txt_filename'],
                "row_count": result['row_count'],
                "company": result['company'],
                "year": result['year']
            })
        else:
            print(f"[API] PDF处理失败: {result['error']}", flush=True)
            return jsonify({"error": result['error']}), 400
            
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/invest_analysis", methods=["POST"])
def invest_analysis():
    try:
        data = request.get_json()
    except Exception as e:
        return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    if not data:
        return jsonify({"error": "请求数据为空"}), 400

    question = data.get("question", "").strip()
    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    try:
        print(f"[API] 收到投资人分析问题: {question}", flush=True)

        # 使用 invest_analysis_stream 收集结果 + thinking_steps（供非H5端使用）
        result = {
            "answer": "",
            "entities": [],
            "file": "",
            "sources": [],
        }
        thinking_steps = []

        for event_type, event_data in bot.invest_analysis_stream(question, verbose=False):
            if event_type == "thinking_step":
                thinking_steps.append({
                    "step": event_data.get("step", 0),
                    "label": event_data.get("label", ""),
                    "status": event_data.get("status", ""),
                    "content": event_data.get("content", "")
                })
            elif event_type == "meta":
                result["entities"] = event_data.get("entities", [])
                result["sources"] = event_data.get("sources", [])
            elif event_type == "chunk":
                result["answer"] += event_data
            elif event_type == "done":
                result["answer"] = event_data.get("answer", result["answer"])
                result["entities"] = event_data.get("entities", result["entities"])
                result["sources"] = event_data.get("sources", result["sources"])

        result["thinking_steps"] = thinking_steps
        print(f"[API] 投资人分析完成, thinking_steps={len(thinking_steps)}", flush=True)
        return jsonify(result)
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        with open("error_log.txt", "a", encoding="utf-8") as f:
            f.write(f"\n=== INVEST ANALYSIS ERROR at {__import__('datetime').datetime.now()} ===\n")
            f.write(f"Question: {question}\n")
            f.write(err_msg)
            f.write("\n")
        return jsonify({"error": str(e)}), 500


@app.route("/api/invest_analysis_stream", methods=["POST"])
def invest_analysis_stream():
    try:
        data = request.get_json()
    except Exception as e:
        return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    if not data:
        return jsonify({"error": "请求数据为空"}), 400

    question = data.get("question", "").strip()
    conversation_id = data.get("conversation_id", "").strip()

    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    def generate():
        try:
            print(f"[API-Stream] 收到投资人分析问题: {question}", flush=True)

            # 关键：发送2KB SSE注释填充，强制Chrome立即开始处理流
            yield ": " + " " * 2048 + "\n\n"

            import json as _json
            import time

            for event_type, event_data in bot.invest_analysis_stream(question, verbose=False):
                if event_type == "thinking_step":
                    step_info = {
                        "type": "thinking_step",
                        "step": event_data.get("step", 0),
                        "label": event_data.get("label", ""),
                        "status": event_data.get("status", ""),
                        "content": event_data.get("content", "")
                    }
                    print(f"[SSE-Invest] thinking_step {step_info['step']} {step_info['status']}", flush=True)
                    yield f"data: {_json.dumps(step_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "meta":
                    meta_info = {
                        "type": "meta",
                        "entities": event_data.get("entities", []),
                        "sources": event_data.get("sources", [])
                    }
                    yield f"data: {_json.dumps(meta_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "chunk":
                    chunk_info = {"type": "chunk", "content": event_data}
                    yield f"data: {_json.dumps(chunk_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "done":
                    done_info = {
                        "type": "done",
                        "answer": event_data.get("answer", ""),
                        "entities": event_data.get("entities", []),
                        "sources": event_data.get("sources", [])
                    }
                    print(f"[SSE-Invest] done", flush=True)
                    yield f"data: {_json.dumps(done_info, ensure_ascii=False)}\n\n"

                    if conversation_id:
                        conversation_manager.add_message(conversation_id, "user", question)
                        conversation_manager.add_message(
                            conversation_id,
                            "bot",
                            event_data.get("answer", ""),
                            extra_data={
                                "thinking": {
                                    "entities": event_data.get("entities", []),
                                }
                            }
                        )

        except Exception as e:
            import traceback as tb
            err_msg = tb.format_exc()
            print(err_msg, flush=True)
            import json as _json
            error_info = {"type": "error", "content": str(e)}
            yield f"data: {_json.dumps(error_info, ensure_ascii=False)}\n\n"

    resp = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )
    return resp


@app.route("/api/invest_analysis_stream_get", methods=["GET"])
def invest_analysis_stream_get():
    """GET 版本的投资人流式分析接口，供前端 EventSource 使用"""
    question = request.args.get("question", "").strip()
    conversation_id = request.args.get("conversation_id", "").strip()

    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    def generate():
        try:
            print(f"[API-Stream-GET] 收到投资人分析问题: {question}", flush=True)

            yield ": " + " " * 2048 + "\n\n"

            import json as _json
            import time

            for event_type, event_data in bot.invest_analysis_stream(question, verbose=False):
                if event_type == "thinking_step":
                    step_info = {
                        "type": "thinking_step",
                        "step": event_data.get("step", 0),
                        "label": event_data.get("label", ""),
                        "status": event_data.get("status", ""),
                        "content": event_data.get("content", "")
                    }
                    yield f"data: {_json.dumps(step_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "meta":
                    meta_info = {
                        "type": "meta",
                        "entities": event_data.get("entities", []),
                        "sources": event_data.get("sources", [])
                    }
                    yield f"data: {_json.dumps(meta_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "chunk":
                    chunk_info = {"type": "chunk", "content": event_data}
                    yield f"data: {_json.dumps(chunk_info, ensure_ascii=False)}\n\n"
                    time.sleep(0)
                elif event_type == "done":
                    done_info = {
                        "type": "done",
                        "answer": event_data.get("answer", ""),
                        "entities": event_data.get("entities", []),
                        "sources": event_data.get("sources", [])
                    }
                    print(f"[SSE-Invest-GET] done, entities={event_data.get('entities', [])}", flush=True)
                    yield f"data: {_json.dumps(done_info, ensure_ascii=False)}\n\n"

                    if conversation_id:
                        conversation_manager.add_message(conversation_id, "user", question)
                        conversation_manager.add_message(
                            conversation_id,
                            "bot",
                            event_data.get("answer", ""),
                            extra_data={
                                "thinking": {
                                    "entities": event_data.get("entities", []),
                                }
                            }
                        )

        except Exception as e:
            import traceback as tb
            err_msg = tb.format_exc()
            print(err_msg, flush=True)
            import json as _json
            error_info = {"type": "error", "content": str(e)}
            yield f"data: {_json.dumps(error_info, ensure_ascii=False)}\n\n"

    resp = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )
    return resp


# PDF 临时存储（token -> 文件路径）
import tempfile
_pdf_temp_store = {}


@app.route("/api/generate_analysis_pdf", methods=["POST"])
def generate_analysis_pdf():
    """生成 PDF 并返回下载 token，供 App 端使用"""
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}

    company = str(data.get("company", "")).strip()
    year = str(data.get("year", "")).strip()
    analysis_text = str(data.get("analysis_text", "")).strip()

    if not company or not year or not analysis_text:
        return jsonify({"error": f"缺少必要参数(company={company}, year={year}, analysis_text={len(analysis_text)} chars)"}), 400

    try:
        print(f"[API] 生成PDF报告(App): {company} {year}", flush=True)
        pdf_bytes = _generate_analysis_pdf(company, year, analysis_text)
        print(f"[API] PDF生成成功 ({len(pdf_bytes)} bytes)", flush=True)

        # 保存到临时文件
        import uuid
        token = str(uuid.uuid4())
        temp_dir = tempfile.gettempdir()
        ascii_company = ""
        for c in company:
            if ord(c) < 128 and (c.isalnum() or c in ('_', '-')):
                ascii_company += c
        if not ascii_company:
            ascii_company = "company"
        temp_filename = f"{ascii_company}_{year}_{token}.pdf"
        temp_path = os.path.join(temp_dir, temp_filename)

        with open(temp_path, "wb") as f:
            f.write(pdf_bytes)

        # 存储映射（5分钟有效）
        _pdf_temp_store[token] = {
            "path": temp_path,
            "filename": f"{ascii_company}_{year}_analysis.pdf",
            "created": __import__('time').time()
        }

        # 清理过期的临时文件（超过10分钟）
        now = __import__('time').time()
        expired_tokens = [k for k, v in _pdf_temp_store.items() if now - v["created"] > 600]
        for t in expired_tokens:
            try:
                os.remove(_pdf_temp_store[t]["path"])
            except Exception:
                pass
            del _pdf_temp_store[t]

        return jsonify({"download_token": token})
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": f"PDF生成失败: {str(e)}"}), 500


@app.route("/api/download_pdf_by_token", methods=["GET"])
def download_pdf_by_token():
    """通过 token 下载已生成的 PDF 文件"""
    token = request.args.get("token", "").strip()
    if not token or token not in _pdf_temp_store:
        return jsonify({"error": "无效或过期的下载链接"}), 404

    info = _pdf_temp_store[token]
    file_path = info["path"]
    filename = info["filename"]

    if not os.path.exists(file_path):
        del _pdf_temp_store[token]
        return jsonify({"error": "文件已过期，请重新生成"}), 404

    try:
        return send_file(
            file_path,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"[API] 下载PDF失败: {e}", flush=True)
        return jsonify({"error": f"下载失败: {str(e)}"}), 500


@app.route("/api/download_analysis_pdf", methods=["GET", "POST"])
def download_analysis_pdf():
    # 从多种来源读取参数（JSON body / form data / URL query）
    data = {}
    try:
        json_data = request.get_json(silent=True)
        if json_data:
            data.update(json_data)
    except Exception:
        pass

    try:
        if request.form:
            for k, v in request.form.items():
                data[k] = v
    except Exception:
        pass

    try:
        if request.args:
            for k, v in request.args.items():
                data[k] = v
    except Exception:
        pass

    # 如果仍然没有数据，尝试手动解析原始 body
    if not data:
        try:
            raw_body = request.get_data(as_text=True)
            if raw_body and raw_body.strip().startswith('{'):
                import json as _json2
                parsed = _json2.loads(raw_body)
                if parsed:
                    data = parsed
        except Exception:
            pass

    print(f"[API] download_analysis_pdf 接收到的数据: {data}", flush=True)

    if not data:
        return jsonify({"error": "未接收到请求数据，请检查请求格式"}), 400

    company = str(data.get("company", "")).strip()
    year = str(data.get("year", "")).strip()
    analysis_text = str(data.get("analysis_text", "")).strip()

    if not company or not year or not analysis_text:
        return jsonify({"error": f"缺少必要参数(company={company}, year={year}, analysis_text={len(analysis_text)} chars)"}), 400

    try:
        print(f"[API] 生成PDF报告: {company} {year}", flush=True)
        pdf_bytes = _generate_analysis_pdf(company, year, analysis_text)

        print(f"[API] PDF生成成功 ({len(pdf_bytes)} bytes)", flush=True)

        import io as _io
        pdf_buffer = _io.BytesIO(pdf_bytes)
        pdf_buffer.seek(0)

        ascii_company = ""
        for c in company:
            if ord(c) < 128 and (c.isalnum() or c in ('_', '-')):
                ascii_company += c
        if not ascii_company:
            ascii_company = "company"

        download_filename = f"{ascii_company}_{year}_analysis.pdf"
        print(f"[API] 发送PDF文件名: {download_filename}", flush=True)

        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=download_filename
        )
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": f"PDF生成失败: {str(e)}"}), 500


# ==================== 对话管理API ====================

@app.route("/api/conversations", methods=["GET"])
def get_conversations():
    try:
        conversations = conversation_manager.get_all_conversations()
        return jsonify({
            "conversations": [
                {
                    "id": c["id"],
                    "title": c["title"],
                    "created_at": c["created_at"],
                    "updated_at": c["updated_at"]
                }
                for c in conversations
            ]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/conversations", methods=["POST"])
def create_conversation():
    try:
        data = request.get_json() or {}
        title = data.get("title", "新对话")
        conversation_id = conversation_manager.create_conversation(title)
        return jsonify({
            "conversation_id": conversation_id,
            "message": "对话创建成功"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/conversations/<conversation_id>", methods=["GET"])
def get_conversation(conversation_id):
    try:
        conversation = conversation_manager.get_conversation(conversation_id)
        if not conversation:
            return jsonify({"error": "对话不存在"}), 404
        
        return jsonify({
            "id": conversation["id"],
            "title": conversation["title"],
            "created_at": conversation["created_at"],
            "updated_at": conversation["updated_at"],
            "messages": conversation["messages"]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/conversations/<conversation_id>", methods=["DELETE"])
def delete_conversation(conversation_id):
    try:
        conversation_manager.delete_conversation(conversation_id)
        return jsonify({"message": "对话删除成功"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/conversations/<conversation_id>/messages", methods=["POST"])
def add_message(conversation_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "请求数据为空"}), 400
        
        role = data.get("role", "").strip()
        content = data.get("content", "").strip()
        
        if not role or not content:
            return jsonify({"error": "角色和内容不能为空"}), 400
        
        success = conversation_manager.add_message(conversation_id, role, content)
        if success:
            return jsonify({"message": "消息添加成功"})
        else:
            return jsonify({"error": "对话不存在"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ==================== 公司对比API（LLM智能分析模式）====================

@app.route("/api/compare_companies", methods=["POST"])
def compare_companies():
    """公司对比分析：使用LLM智能提取财务指标"""
    try:
        data = request.get_json()
    except Exception as e:
        return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    if not data:
        return jsonify({"error": "请求数据为空"}), 400

    companies = data.get("companies", [])
    year = str(data.get("year", "")).strip()
    
    if not companies or len(companies) == 0:
        return jsonify({"error": "请至少选择一家公司"}), 400

    try:
        print(f"[API] 公司对比(LLM模式): {companies}, 年份: {year}", flush=True)
        
        result = {
            "companies": [],
            "radar_data": {},
            "gross_margin_trend": {},
            "years": [],
            "calculations": {},
            "raw_contexts": {},
            "analysis_mode": "llm",
            "not_found": []
        }
        
        for company in companies:
            print(f"[API] 正在LLM分析: {company} ...", flush=True)
            
            analysis = bot.llm_financial_analysis(company, year=year, verbose=True)
            
            if analysis.get("error"):
                print(f"[API] {company} 分析失败: {analysis['error']}", flush=True)
                result["not_found"].append({"company": company, "reason": analysis["error"]})
                result["raw_contexts"][company] = analysis.get("raw_context", [])
                continue
            
            result["companies"].append(company)
            result["radar_data"][company] = analysis.get("radar_data", {})
            result["gross_margin_trend"][company] = analysis.get("gross_margin_trend", [])
            result["calculations"][company] = analysis.get("calculations", {})
            result["raw_contexts"][company] = analysis.get("raw_context", [])
            
            trend = analysis.get("gross_margin_trend", [])
            if trend and len(trend) > 1:
                entities = [company]
                if year:
                    entities.append(year)
                txt_path = bot._find_txt_file(entities)
                if txt_path:
                    fname = os.path.basename(txt_path)
                    parts = fname.replace(".txt", "").split("__")
                    file_year = ""
                    for p in parts:
                        m = re.search(r"(\d{4})", p)
                        if m:
                            file_year = m.group(1)
                            break
                    if file_year:
                        y = int(file_year)
                        years = [str(y - i) for i in range(len(trend) - 1, -1, -1)]
                        if not result["years"] or len(years) > len(result["years"]):
                            result["years"] = years
        
        print(f"[API] 最终结果: companies={result['companies']}", flush=True)
        for c in result["companies"]:
            print(f"[API]   {c}: radar={list(result['radar_data'][c].keys())}, calc={list(result['calculations'][c].keys())}", flush=True)
        
        return jsonify(result)
    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": str(e)}), 500


# ==================== 职业分析API（LLM智能分析模式）====================

@app.route("/api/career_analysis", methods=["GET", "POST"])
def career_analysis():
    """职业分析：使用LLM智能提取财务指标并生成入职建议"""
    data = {}
    try:
        json_data = request.get_json(silent=True)
        if json_data:
            data.update(json_data)
    except Exception:
        pass

    try:
        if request.form:
            for k, v in request.form.items():
                data[k] = v
    except Exception:
        pass

    try:
        if request.args:
            for k, v in request.args.items():
                data[k] = v
    except Exception:
        pass

    if not data:
        try:
            raw_body = request.get_data(as_text=True)
            if raw_body and raw_body.strip().startswith('{'):
                import json as _json2
                parsed = _json2.loads(raw_body)
                if parsed:
                    data = parsed
        except Exception:
            pass

    print(f"[API] career_analysis 接收到的数据: {data}", flush=True)

    if not data:
        return jsonify({"error": "未接收到请求数据，请检查请求格式"}), 400

    companies = data.get("companies", [])
    if isinstance(companies, str):
        import json as _json3
        try:
            companies = _json3.loads(companies)
        except Exception:
            companies = [c.strip() for c in companies.replace('，', ',').split(',') if c.strip()]
    year = str(data.get("year", "")).strip()

    if not companies or len(companies) == 0:
        return jsonify({"error": "请至少选择一家公司"}), 400

    try:
        print(f"[API] 职业分析(LLM模式): {companies}, 年份: {year}", flush=True)

        companies_data = []

        for company in companies:
            print(f"[API] 职业分析: 正在LLM分析 {company} ...", flush=True)
            
            analysis = bot.llm_financial_analysis(company, year=year, verbose=True)
            
            if analysis.get("error"):
                print(f"[API] 职业分析: {company} 失败: {analysis['error']}", flush=True)
                continue
            
            companies_data.append({
                "company": company,
                "radar_data": analysis.get("radar_data", {}),
                "calculations": analysis.get("calculations", {}),
                "raw_context": analysis.get("raw_context", [])
            })

        if len(companies_data) == 0:
            return jsonify({"error": "未能提取到任何公司的财务数据"}), 400

        # 使用 career_analysis prompt 生成入职建议
        analysis_result = bot.career_analysis(companies_data, verbose=False)
        answer_text = analysis_result.get("answer", "")
        
        # 解析人格得分 JSON
        personality_scores = {}
        import json as _json
        try:
            json_match = re.search(r'PERSONALITY_SCORES_JSON_START\s*(\{.*?\})\s*PERSONALITY_SCORES_JSON_END', answer_text, re.DOTALL)
            if json_match:
                personality_scores = _json.loads(json_match.group(1))
                print(f"[API] 成功解析人格得分: {len(personality_scores)} 种人格", flush=True)
            else:
                print(f"[API] 未找到人格得分JSON，使用默认值", flush=True)
                # 生成默认得分
                mbti_types = ["ISTJ","ISFJ","INFJ","INTJ","ISTP","ISFP","INFP","INTP",
                              "ESTP","ESFP","ENFP","ENTP","ESTJ","ESFJ","ENFJ","ENTJ"]
                for mbti in mbti_types:
                    personality_scores[mbti] = {c["company"]: 50 for c in companies_data}
        except Exception as e:
            print(f"[API] 解析人格得分失败: {e}", flush=True)
            mbti_types = ["ISTJ","ISFJ","INFJ","INTJ","ISTP","ISFP","INFP","INTP",
                          "ESTP","ESFP","ENFP","ENTP","ESTJ","ESFJ","ENFJ","ENTJ"]
            for mbti in mbti_types:
                personality_scores[mbti] = {c["company"]: 50 for c in companies_data}

        return jsonify({
            "companies": [c["company"] for c in companies_data],
            "radar_data": {c["company"]: c["radar_data"] for c in companies_data},
            "calculations": {c["company"]: c["calculations"] for c in companies_data},
            "raw_contexts": {c["company"]: c.get("raw_context", []) for c in companies_data},
            "career_analysis": answer_text,
            "personality_scores": personality_scores
        })

    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": str(e)}), 500


# ==================== 职业分析Excel下载API ====================

@app.route("/api/download_career_excel", methods=["GET", "POST"])
def download_career_excel():
    """生成并下载职业分析Excel表格（16种MBTI人格类型 x 公司推荐得分）"""
    data = {}
    try:
        json_data = request.get_json(silent=True)
        if json_data:
            data.update(json_data)
    except Exception:
        pass

    try:
        if request.form:
            for k, v in request.form.items():
                data[k] = v
    except Exception:
        pass

    try:
        if request.args:
            for k, v in request.args.items():
                data[k] = v
    except Exception:
        pass

    if not data:
        try:
            raw_body = request.get_data(as_text=True)
            if raw_body and raw_body.strip().startswith('{'):
                import json as _json2
                parsed = _json2.loads(raw_body)
                if parsed:
                    data = parsed
        except Exception:
            pass

    print(f"[API] download_career_excel 接收到的数据: {data}", flush=True)

    if not data:
        return jsonify({"error": "未接收到请求数据，请检查请求格式"}), 400

    companies = data.get("companies", [])
    if isinstance(companies, str):
        import json as _json3
        try:
            companies = _json3.loads(companies)
        except Exception:
            companies = [c.strip() for c in companies.replace('，', ',').split(',') if c.strip()]

    personality_scores = data.get("personality_scores", {})
    if isinstance(personality_scores, str):
        import json as _json4
        try:
            personality_scores = _json4.loads(personality_scores)
        except Exception:
            personality_scores = {}

    if not companies or not personality_scores:
        return jsonify({"error": "缺少必要数据"}), 400

    try:
        import io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            os.system("pip install openpyxl -q")
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "人格匹配度评分"

        # 标题行
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 表头：人格类型 | 公司1 | 公司2 | ...
        headers = ["MBTI人格类型"] + companies
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行
        mbti_order = ["ISTJ","ISFJ","INFJ","INTJ","ISTP","ISFP","INFP","INTP",
                      "ESTP","ESFP","ENFP","ENTP","ESTJ","ESFJ","ENFJ","ENTJ"]
        mbti_names = {
            "ISTJ": "ISTJ 务实管家", "ISFJ": "ISFJ 守护者", "INFJ": "INFJ 理想主义者",
            "INTJ": "INTJ 战略家", "ISTP": "ISTP 鉴赏家", "ISFP": "ISFP 艺术家",
            "INFP": "INFP 调停者", "INTP": "INTP 思想家", "ESTP": "ESTP 企业家",
            "ESFP": "ESFP 表演者", "ENFP": "ENFP 竞选者", "ENTP": "ENTP 辩论家",
            "ESTJ": "ESTJ 总经理", "ESFJ": "ESFJ 执政官", "ENFJ": "ENFJ 主人公",
            "ENTJ": "ENTJ 指挥官"
        }

        row_colors = [
            PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]

        for row_idx, mbti in enumerate(mbti_order, 2):
            scores = personality_scores.get(mbti, {})
            cell = ws.cell(row=row_idx, column=1, value=mbti_names.get(mbti, mbti))
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_colors[row_idx % 2]
            
            for col_idx, company in enumerate(companies, 2):
                score = scores.get(company, 50)
                cell = ws.cell(row=row_idx, column=col_idx, value=score)
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = row_colors[row_idx % 2]
                
                # 高分绿色、低分红色
                if score >= 80:
                    cell.font = Font(color="0D652D", bold=True)
                elif score >= 60:
                    cell.font = Font(color="196138")
                elif score < 40:
                    cell.font = Font(color="CC0000")

        # 设置列宽
        ws.column_dimensions['A'].width = 22
        for col in range(2, len(companies) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # 冻结首行
        ws.freeze_panes = "B2"

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"career_analysis_{'_'.join(companies[:3])}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback as tb
        err_msg = tb.format_exc()
        print(err_msg, flush=True)
        return jsonify({"error": f"Excel生成失败: {str(e)}"}), 500


@app.route("/api/generate_career_excel", methods=["POST"])
def generate_career_excel():
    """生成 Excel 并返回下载 token，供 App 端使用"""
    data = {}
    try:
        json_data = request.get_json(silent=True)
        if json_data:
            data.update(json_data)
    except Exception:
        pass

    if not data:
        try:
            raw_body = request.get_data(as_text=True)
            if raw_body and raw_body.strip().startswith('{'):
                import json as _json5
                parsed = _json5.loads(raw_body)
                if parsed:
                    data = parsed
        except Exception:
            pass

    companies = data.get("companies", [])
    personality_scores = data.get("personality_scores", {})

    if not companies or not personality_scores:
        return jsonify({"error": "缺少必要数据"}), 400

    try:
        # 复用 download_career_excel 的逻辑生成 Excel
        import io as _io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            os.system("pip install openpyxl -q")
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "人格匹配度评分"

        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["MBTI人格类型"] + companies
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        mbti_order = ["ISTJ","ISFJ","INFJ","INTJ","ISTP","ISFP","INFP","INTP",
                      "ESTP","ESFP","ENFP","ENTP","ESTJ","ESFJ","ENFJ","ENTJ"]
        mbti_names = {
            "ISTJ": "ISTJ 务实管家", "ISFJ": "ISFJ 守护者", "INFJ": "INFJ 理想主义者",
            "INTJ": "INTJ 战略家", "ISTP": "ISTP 鉴赏家", "ISFP": "ISFP 艺术家",
            "INFP": "INFP 调停者", "INTP": "INTP 逻辑学家", "ESTP": "ESTP 企业家",
            "ESFP": "ESFP 表演者", "ENFP": "ENFP 竞选者", "ENTP": "ENTP 辩论家",
            "ESTJ": "ESTJ 总经理", "ESFJ": "ESFJ 执政官", "ENFJ": "ENFJ 主人公",
            "ENTJ": "ENTJ 指挥官"
        }

        for row_idx, mbti in enumerate(mbti_order, 2):
            cell = ws.cell(row=row_idx, column=1, value=mbti_names.get(mbti, mbti))
            cell.font = Font(bold=True, size=10)
            cell.alignment = center_align
            cell.border = thin_border

            scores_for_type = personality_scores.get(mbti, {})
            for col_idx, company in enumerate(companies, 2):
                val = scores_for_type.get(company, 0)
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = 0
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin_border
                if val >= 80:
                    cell.fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
                elif val >= 60:
                    cell.fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")

        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(companies) + 2):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        output = _io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 保存到临时文件
        import uuid
        token = str(uuid.uuid4())
        temp_dir = tempfile.gettempdir()
        temp_filename = f"career_{token}.xlsx"
        temp_path = os.path.join(temp_dir, temp_filename)

        with open(temp_path, "wb") as f:
            f.write(output.getvalue())

        _pdf_temp_store[token] = {
            "path": temp_path,
            "filename": f"career_analysis_{'_'.join(companies[:3])}.xlsx",
            "created": __import__('time').time()
        }

        return jsonify({"download_token": token})
    except Exception as e:
        import traceback as tb
        print(tb.format_exc(), flush=True)
        return jsonify({"error": f"Excel生成失败: {str(e)}"}), 500


@app.route("/api/download_excel_by_token", methods=["GET"])
def download_excel_by_token():
    """通过 token 下载已生成的 Excel 文件"""
    token = request.args.get("token", "").strip()
    if not token or token not in _pdf_temp_store:
        return jsonify({"error": "无效或过期的下载链接"}), 404

    info = _pdf_temp_store[token]
    file_path = info["path"]
    filename = info["filename"]

    if not os.path.exists(file_path):
        del _pdf_temp_store[token]
        return jsonify({"error": "文件已过期，请重新生成"}), 404

    try:
        return send_file(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"[API] 下载Excel失败: {e}", flush=True)
        return jsonify({"error": f"下载失败: {str(e)}"}), 500


if __name__ == "__main__":
    # 自定义请求处理器：禁用所有缓冲，确保 SSE 数据立即发送
    class NoDelayRequestHandler(WSGIRequestHandler):
        # 关键：禁用 wfile 写入缓冲，否则 Werkzeug 会攒数据再发
        wbufsize = 0

        def handle(self):
            try:
                self.connection.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)
            except Exception:
                pass
            super().handle()

    print()
    print("=" * 50)
    print("  ChatFinance API Server (LLM Analysis Mode)")
    print("  后端: http://localhost:5000/api/chat")
    print("  前端: http://localhost:5000")
    print("=" * 50)
    print()
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True, request_handler=NoDelayRequestHandler)
